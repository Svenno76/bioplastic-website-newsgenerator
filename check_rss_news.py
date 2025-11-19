#!/usr/bin/env python3
"""
Check company websites for RSS feeds and news sections.
Adds two new columns to companies.xlsx:
- RSS Feed URL
- News Section URL
"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import time
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def normalize_url(url):
    """Add https:// if missing"""
    if not url or pd.isna(url):
        return None
    url = str(url).strip()
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    return url

def find_rss_feed(url, soup, session):
    """
    Find RSS feed URLs on the page.
    Checks:
    1. <link> tags with type="application/rss+xml"
    2. Common RSS feed paths
    """
    rss_feeds = []

    # Method 1: Look for RSS link tags in HTML
    link_tags = soup.find_all('link', {'type': re.compile(r'application/(rss|atom)\+xml', re.I)})
    for tag in link_tags:
        href = tag.get('href')
        if href:
            full_url = urljoin(url, href)
            rss_feeds.append(full_url)

    # Method 2: Check common RSS feed paths
    common_paths = [
        '/feed',
        '/rss',
        '/blog/feed',
        '/news/feed',
        '/feed/',
        '/rss/',
        '/atom.xml',
        '/rss.xml',
        '/feed.xml'
    ]

    base_url = f"{urlparse(url).scheme}://{urlparse(url).netloc}"

    for path in common_paths:
        test_url = base_url + path
        try:
            response = session.head(test_url, timeout=5, allow_redirects=True)
            if response.status_code == 200:
                content_type = response.headers.get('content-type', '').lower()
                if 'xml' in content_type or 'rss' in content_type or 'atom' in content_type:
                    rss_feeds.append(test_url)
        except:
            pass

    # Return the first RSS feed found
    return rss_feeds[0] if rss_feeds else None

def find_news_section(url, soup):
    """
    Find news/press/media section URLs on the page.
    Looks for links containing keywords like news, press, media, blog.
    """
    news_keywords = ['news', 'press', 'media', 'blog', 'press-release', 'newsroom',
                     'announcements', 'updates', 'press-releases']

    news_urls = []

    # Find all links
    links = soup.find_all('a', href=True)

    for link in links:
        href = link.get('href', '')
        text = link.get_text().lower().strip()

        # Check if the link text or href contains news-related keywords
        if any(keyword in href.lower() for keyword in news_keywords) or \
           any(keyword in text for keyword in news_keywords):
            full_url = urljoin(url, href)

            # Filter out external links and irrelevant paths
            if urlparse(full_url).netloc == urlparse(url).netloc:
                # Prefer paths that are likely news sections (not individual articles)
                if not re.search(r'/\d{4}/', full_url):  # Avoid date-based URLs
                    news_urls.append(full_url)

    # Return the first news section URL found (prioritize shorter paths)
    if news_urls:
        news_urls.sort(key=len)
        return news_urls[0]

    return None

def check_website(url):
    """
    Check a website for RSS feed and news section.
    Returns: (rss_url, news_url)
    """
    url = normalize_url(url)
    if not url:
        return None, None

    try:
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

        print(f"  Checking: {url}")
        response = session.get(url, timeout=10, allow_redirects=True)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')

        # Find RSS feed
        rss_url = find_rss_feed(url, soup, session)

        # Find news section
        news_url = find_news_section(url, soup)

        print(f"    RSS: {rss_url if rss_url else 'Not found'}")
        print(f"    News: {news_url if news_url else 'Not found'}")

        return rss_url, news_url

    except Exception as e:
        print(f"    Error: {str(e)}")
        return None, None

def format_excel_file(filename):
    """Format the Excel file with clickable URLs and optimized column widths"""
    wb = load_workbook(filename)
    ws = wb.active

    # Define column widths
    column_widths = {
        'Company': 25,
        'Type': 20,
        'Country': 15,
        'Webpage': 40,
        'Description': 70,
        'Primary Materials': 50,
        'Market Segments': 50,
        'Status': 12,
        'Publicly Listed': 15,
        'Stock Ticker': 15,
        'Date Added': 15,
        'RSS Feed URL': 50,
        'News Section URL': 50
    }

    # Get header row
    headers = [cell.value for cell in ws[1]]

    # Set column widths and format
    for idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(idx)

        # Set width
        if header in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[header]

        # Set text wrapping for long text columns
        if header in ['Description', 'Primary Materials', 'Market Segments', 'Headline']:
            for row in range(2, ws.max_row + 1):
                ws[f'{col_letter}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    # Make URLs clickable
    url_columns = ['Webpage', 'RSS Feed URL', 'News Section URL']
    for col_name in url_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            col_letter = get_column_letter(col_idx)

            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if cell.value and str(cell.value).strip():
                    url = str(cell.value).strip()
                    # Add https:// if missing
                    if not url.startswith(('http://', 'https://')):
                        url = 'https://' + url
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")

    wb.save(filename)
    print(f"\n✅ Excel file formatted with clickable URLs and optimized column widths")

def main():
    print("=" * 60)
    print("RSS Feed and News Section Checker")
    print("=" * 60)

    # Read the companies file
    print("\n📖 Reading companies.xlsx...")
    df = pd.read_excel('companies.xlsx')

    print(f"Found {len(df)} companies")

    # Add new columns if they don't exist
    if 'RSS Feed URL' not in df.columns:
        df['RSS Feed URL'] = None
    if 'News Section URL' not in df.columns:
        df['News Section URL'] = None

    # Check each company's website
    print("\n🔍 Checking websites for RSS feeds and news sections...\n")

    for idx, row in df.iterrows():
        company = row['Company']
        webpage = row.get('Webpage')

        print(f"\n[{idx+1}/{len(df)}] {company}")

        if pd.isna(webpage) or not webpage:
            print("  No webpage - skipping")
            continue

        # Check website
        rss_url, news_url = check_website(webpage)

        # Update dataframe
        df.at[idx, 'RSS Feed URL'] = rss_url
        df.at[idx, 'News Section URL'] = news_url

        # Rate limiting
        time.sleep(2)

    # Save updated file
    print("\n" + "=" * 60)
    print("💾 Saving results to companies.xlsx...")

    # Create backup
    import shutil
    shutil.copy('companies.xlsx', 'companies_backup.xlsx')
    print("✅ Backup created: companies_backup.xlsx")

    # Save to Excel
    df.to_excel('companies.xlsx', index=False, engine='openpyxl')
    print("✅ Data saved to companies.xlsx")

    # Format Excel file
    format_excel_file('companies.xlsx')

    # Print summary
    rss_count = df['RSS Feed URL'].notna().sum()
    news_count = df['News Section URL'].notna().sum()

    print("\n" + "=" * 60)
    print("📊 Summary")
    print("=" * 60)
    print(f"Total companies checked: {len(df)}")
    print(f"RSS feeds found: {rss_count} ({rss_count/len(df)*100:.1f}%)")
    print(f"News sections found: {news_count} ({news_count/len(df)*100:.1f}%)")
    print("\n✅ Complete!")
    print("=" * 60)

if __name__ == '__main__':
    main()
