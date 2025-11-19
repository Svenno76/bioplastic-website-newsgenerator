#!/usr/bin/env python3
"""
Company Enrichment Skill
Automatically researches and fills in missing company information using Perplexity AI
"""

import sys
import os
import json
import re
from datetime import datetime
from pathlib import Path
import pandas as pd
import requests
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Add parent directory to path to import config
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

try:
    from config import Config
except ImportError:
    print("❌ Error: Could not import config.py")
    print("Make sure config.py exists in the project root")
    sys.exit(1)

# Company type categories
VALID_COMPANY_TYPES = [
    "Bioplastic Producer",
    "Compounder",
    "Converter",
    "Technology Company",
    "Equipment Manufacturer",
    "Additive Producer",
    "Testing/Certification Company",
    "Distributor/Trader",
    "Recycling Company",
    "Waste Management"
]

# Valid status values
VALID_STATUSES = ["Active", "Acquired", "Defunct", "Unknown"]


def print_header():
    """Print skill header"""
    print("\n" + "=" * 70)
    print("🏢 COMPANY ENRICHMENT SKILL")
    print("=" * 70 + "\n")


def load_companies(file_path):
    """Load companies from Excel file"""
    print(f"📂 Loading {file_path}...")

    if not os.path.exists(file_path):
        print(f"❌ Error: {file_path} not found")
        return None

    try:
        df = pd.read_excel(file_path)
        print(f"  ✓ Loaded {len(df)} companies")
        return df
    except Exception as e:
        print(f"❌ Error loading file: {e}")
        return None


def identify_incomplete_records(df):
    """Find companies with missing data"""
    # Fields to check for completeness
    fields_to_check = ['Type', 'Country', 'Description', 'Primary Materials',
                       'Market Segments', 'Status', 'Publicly Listed']

    # Add social media fields if they exist
    social_media_fields = ['Twitter', 'LinkedIn', 'YouTube', 'Instagram']
    for field in social_media_fields:
        if field in df.columns:
            fields_to_check.append(field)

    # Find rows with at least one empty field
    incomplete_mask = df[fields_to_check].isnull().any(axis=1) | \
                      (df[fields_to_check] == '').any(axis=1)

    incomplete_df = df[incomplete_mask].copy()

    if len(incomplete_df) > 0:
        print(f"  ⚠️  Found {len(incomplete_df)} companies with incomplete data")
        for idx, row in incomplete_df.iterrows():
            empty_fields = [field for field in fields_to_check
                          if pd.isna(row[field]) or row[field] == '']
            print(f"     - {row['Company']}: missing {', '.join(empty_fields)}")
    else:
        print("  ✓ All companies have complete data")

    return incomplete_df


def delete_unknown_companies(df):
    """Delete companies with Type = 'Unknown'"""
    unknown_mask = df['Type'] == 'Unknown'
    unknown_companies = df[unknown_mask]['Company'].tolist()

    if len(unknown_companies) > 0:
        print(f"\n🗑️  Deleting {len(unknown_companies)} companies with Type = 'Unknown':")
        for company in unknown_companies:
            print(f"     - {company}")

        # Remove unknown companies
        df = df[~unknown_mask].copy()
        df = df.reset_index(drop=True)
        print(f"  ✓ Removed {len(unknown_companies)} companies")
        print(f"  ✓ Remaining companies: {len(df)}")
    else:
        print("\n  ✓ No companies with Type = 'Unknown' found")

    return df


def create_backup(file_path):
    """Create backup of companies file"""
    backup_path = file_path.replace('.xlsx', '_backup.xlsx')
    try:
        df = pd.read_excel(file_path)
        df.to_excel(backup_path, index=False)
        print(f"  ✓ Created backup: {backup_path}")
        return True
    except Exception as e:
        print(f"  ⚠️  Could not create backup: {e}")
        return False


def research_company(company_name, company_website=None):
    """
    Research a company using Perplexity AI

    Args:
        company_name: Name of the company to research
        company_website: Optional known website for validation

    Returns:
        dict: Company information or None if failed
    """
    print(f"\n📡 Researching: {company_name}")

    # Build the research query
    query = f"""
Research the bioplastic company "{company_name}" and provide the following information in JSON format:

{{
  "Type": "One of: {', '.join(VALID_COMPANY_TYPES)}",
  "Country": "Headquarters country (full name)",
  "Description": "2-3 sentence overview of the company, its products, and specialties",
  "PrimaryMaterials": "Specific bioplastics they produce/use (e.g., PLA, PHA, PBS, starch-based, bio-PE, etc.)",
  "MarketSegments": "Industries served (e.g., packaging, agriculture, automotive, medical, textiles, etc.)",
  "Status": "One of: Active, Acquired, Defunct, Unknown",
  "PubliclyListed": "Yes or No - is the company publicly traded on a stock exchange?",
  "StockTicker": "Stock ticker symbol (e.g., NASDAQ:DNMR, NYSE:AMCR) if publicly listed, otherwise leave blank",
  "Webpage": "Official company website URL (validate and correct if needed)",
  "Twitter": "Official company Twitter/X handle (e.g., @CompanyName or full URL like https://twitter.com/CompanyName), or leave blank if not found",
  "LinkedIn": "Official company LinkedIn page URL (e.g., https://www.linkedin.com/company/company-name), or leave blank if not found",
  "YouTube": "Official company YouTube channel URL (e.g., https://www.youtube.com/@ChannelName), or leave blank if not found",
  "Instagram": "Official company Instagram profile URL (e.g., https://www.instagram.com/username), or leave blank if not found"
}}

IMPORTANT:
- Type MUST be exactly one of the listed categories
- Status MUST be one of: Active, Acquired, Defunct, Unknown
- PubliclyListed MUST be "Yes" or "No"
- StockTicker should include exchange prefix if known (e.g., NASDAQ:DNMR, NYSE:AMCR, TSE:4118)
- If not publicly listed, leave StockTicker blank
- Country should be the full name (e.g., "United States" not "USA")
- Description should be concise (50-150 words)
- Focus on bioplastic-related activities
- For social media: Only include verified official company accounts
- Social media URLs should be complete and valid (not shortened links)
- If information is uncertain, leave social media field blank rather than guessing
- If the company does not have a particular social media account, leave that field blank

Return ONLY valid JSON, no markdown formatting or explanations.
"""

    if company_website:
        query += f"\n\nKnown website: {company_website}"

    # Prepare API request
    headers = {
        "Authorization": f"Bearer {Config.PERPLEXITY_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": Config.DEFAULT_MODEL,
        "messages": [
            {
                "role": "system",
                "content": "You are a research assistant specializing in the bioplastic industry. Provide accurate, structured information about companies in JSON format. Focus on finding verified official social media accounts."
            },
            {
                "role": "user",
                "content": query
            }
        ],
        "max_tokens": Config.MAX_TOKENS,
        "temperature": 0.2,  # Lower temperature for more factual responses
        "return_citations": True,
        "stream": False
    }

    try:
        response = requests.post(
            Config.PERPLEXITY_API_URL,
            headers=headers,
            json=payload,
            timeout=30
        )
        response.raise_for_status()

        result = response.json()
        content = result['choices'][0]['message']['content']

        # Clean up markdown formatting if present
        content = content.strip()
        if content.startswith('```json'):
            content = content[7:]
        if content.startswith('```'):
            content = content[3:]
        if content.endswith('```'):
            content = content[:-3]
        content = content.strip()

        # Parse JSON
        company_data = json.loads(content)

        print(f"  ✓ Research complete")
        return company_data

    except requests.exceptions.RequestException as e:
        print(f"  ❌ API request failed: {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"  ❌ Failed to parse JSON response: {e}")
        print(f"  Raw response: {content[:200]}...")
        return None
    except Exception as e:
        print(f"  ❌ Unexpected error: {e}")
        return None


def normalize_social_media_url(url, platform):
    """
    Normalize social media URLs to standard formats.

    Args:
        url: Social media URL or handle
        platform: Platform name (Twitter, LinkedIn, YouTube, Instagram)

    Returns:
        str: Normalized full URL
    """
    url = url.strip()

    if platform == 'Twitter':
        # Handle @username or twitter.com/username formats
        if url.startswith('@'):
            return f'https://twitter.com/{url[1:]}'
        elif 'twitter.com' in url or 'x.com' in url:
            if not url.startswith('http'):
                return f'https://{url}'
            return url
        else:
            # Assume it's a handle
            return f'https://twitter.com/{url}'

    elif platform == 'LinkedIn':
        # Handle LinkedIn URLs
        if 'linkedin.com' not in url:
            return f'https://www.linkedin.com/company/{url}'
        elif not url.startswith('http'):
            return f'https://{url}'
        return url

    elif platform == 'YouTube':
        # Handle YouTube URLs
        if 'youtube.com' not in url and 'youtu.be' not in url:
            return f'https://www.youtube.com/@{url}'
        elif not url.startswith('http'):
            return f'https://{url}'
        return url

    elif platform == 'Instagram':
        # Handle Instagram URLs
        if 'instagram.com' not in url:
            return f'https://www.instagram.com/{url}/'
        elif not url.startswith('http'):
            return f'https://{url}'
        return url

    return url


def is_valid_url(url):
    """
    Validate if URL has proper format.

    Args:
        url: URL to validate

    Returns:
        bool: True if valid URL format
    """
    try:
        parsed = urlparse(url)
        return bool(parsed.scheme and parsed.netloc)
    except:
        return False


def validate_company_data(data, company_name):
    """
    Validate and clean company data

    Args:
        data: Dict of company information
        company_name: Name of the company being validated

    Returns:
        dict: Validated data
    """
    validated = {}

    # Validate Type
    company_type = data.get('Type', 'Unknown')
    if company_type not in VALID_COMPANY_TYPES:
        print(f"  ⚠️  Invalid type '{company_type}', setting to Unknown")
        # Try to find closest match
        for valid_type in VALID_COMPANY_TYPES:
            if valid_type.lower() in company_type.lower() or company_type.lower() in valid_type.lower():
                company_type = valid_type
                print(f"  ✓ Matched to '{valid_type}'")
                break
        else:
            company_type = "Unknown"
    validated['Type'] = company_type

    # Validate Country
    country = data.get('Country', 'Unknown')
    if not country or country.strip() == '':
        country = 'Unknown'
    validated['Country'] = country.strip()

    # Validate Description
    description = data.get('Description', '')
    if description and len(description) > 500:
        description = description[:497] + '...'
    validated['Description'] = description.strip()

    # Validate Primary Materials
    materials = data.get('PrimaryMaterials', 'Unknown')
    if isinstance(materials, list):
        materials = ', '.join(materials)
    validated['Primary Materials'] = materials.strip() if isinstance(materials, str) else str(materials)

    # Validate Market Segments
    segments = data.get('MarketSegments', 'Unknown')
    if isinstance(segments, list):
        segments = ', '.join(segments)
    validated['Market Segments'] = segments.strip() if isinstance(segments, str) else str(segments)

    # Validate Status
    status = data.get('Status', 'Unknown')
    if status not in VALID_STATUSES:
        print(f"  ⚠️  Invalid status '{status}', setting to Unknown")
        status = 'Unknown'
    validated['Status'] = status

    # Validate Webpage
    webpage = data.get('Webpage', '')
    if webpage:
        # Add https:// if missing
        if not webpage.startswith(('http://', 'https://')):
            webpage = 'https://' + webpage

        # Basic URL validation
        try:
            parsed = urlparse(webpage)
            if parsed.scheme and parsed.netloc:
                validated['Webpage'] = webpage
            else:
                print(f"  ⚠️  Invalid webpage URL: {webpage}")
                validated['Webpage'] = ''
        except:
            print(f"  ⚠️  Could not parse webpage URL: {webpage}")
            validated['Webpage'] = ''
    else:
        validated['Webpage'] = ''

    # Validate Publicly Listed
    publicly_listed = data.get('PubliclyListed', 'No').strip()
    if publicly_listed.lower() in ['yes', 'y', 'true']:
        publicly_listed = 'Yes'
    elif publicly_listed.lower() in ['no', 'n', 'false']:
        publicly_listed = 'No'
    else:
        print(f"  ⚠️  Invalid PubliclyListed value '{publicly_listed}', setting to No")
        publicly_listed = 'No'
    validated['Publicly Listed'] = publicly_listed

    # Validate Stock Ticker
    stock_ticker = data.get('StockTicker', '').strip()
    if publicly_listed == 'No':
        stock_ticker = ''  # Clear ticker if not publicly listed
    validated['Stock Ticker'] = stock_ticker

    # Validate Social Media URLs
    social_media_fields = {
        'Twitter': 'Twitter',
        'LinkedIn': 'LinkedIn',
        'YouTube': 'YouTube',
        'Instagram': 'Instagram'
    }

    for field_key, field_name in social_media_fields.items():
        social_url = data.get(field_key, '').strip()

        if social_url:
            # Normalize social media URLs
            social_url = normalize_social_media_url(social_url, field_name)

            # Validate if it's a valid URL format
            if social_url and is_valid_url(social_url):
                validated[field_name] = social_url
            else:
                print(f"  ⚠️  Invalid {field_name} URL: {data.get(field_key, '')}")
                validated[field_name] = ''
        else:
            validated[field_name] = ''

    # Add timestamp
    validated['Date Added'] = datetime.now().strftime('%Y-%m-%d')

    return validated


def enrich_company(df, idx, company_name):
    """
    Enrich a single company's data

    Args:
        df: DataFrame with company data
        idx: Index of company to enrich
        company_name: Name of the company

    Returns:
        bool: True if successful, False otherwise
    """
    # Get existing webpage if available
    existing_webpage = df.at[idx, 'Webpage']
    if pd.isna(existing_webpage):
        existing_webpage = None

    # Research the company
    company_data = research_company(company_name, existing_webpage)

    if not company_data:
        print(f"  ❌ Failed to research {company_name}")
        return False

    # Validate data
    validated_data = validate_company_data(company_data, company_name)

    # Update DataFrame - only fill empty fields
    for field, value in validated_data.items():
        if field in df.columns:
            # Only update if current value is empty/null
            current_value = df.at[idx, field]
            if pd.isna(current_value) or current_value == '':
                df.at[idx, field] = value
                print(f"  ✓ {field}: {value}")
            else:
                print(f"  ⏭️  {field}: keeping existing value")

    return True


def format_companies_excel(file_path):
    """
    Format companies Excel file with clickable URLs, proper widths, and text wrapping
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Define column widths and wrap settings
        column_config = {
            'A': ('Company', 25, False),
            'B': ('Type', 20, False),
            'C': ('Country', 15, False),
            'D': ('Webpage', 40, False),
            'E': ('Description', 70, True),
            'F': ('Primary Materials', 50, True),
            'G': ('Market Segments', 50, True),
            'H': ('Status', 12, False),
            'I': ('Publicly Listed', 15, False),
            'J': ('Stock Ticker', 15, False),
            'K': ('Twitter', 35, False),
            'L': ('LinkedIn', 35, False),
            'M': ('YouTube', 35, False),
            'N': ('Instagram', 35, False),
            'O': ('Date Added', 15, False),
        }

        # Set column widths and wrap text
        for col_letter, (col_name, width, wrap) in column_config.items():
            ws.column_dimensions[col_letter].width = width

            # Apply to all cells in column (skip header)
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if wrap:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Make URLs clickable for multiple columns
        url_columns = {
            'D': 'Webpage',
            'K': 'Twitter',
            'L': 'LinkedIn',
            'M': 'YouTube',
            'N': 'Instagram'
        }

        for col_letter, col_name in url_columns.items():
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if cell.value:
                    url = str(cell.value).strip()

                    # Normalize URL format
                    if url and not url.startswith(('http://', 'https://')):
                        if url.startswith('www.') or '.' in url:
                            url = 'https://' + url

                    # Make clickable if valid URL
                    if url.startswith('http'):
                        cell.hyperlink = url
                        cell.style = 'Hyperlink'

        wb.save(file_path)
    except Exception as e:
        print(f"  ⚠️  Could not format Excel file: {e}")


def save_results(df, file_path):
    """Save enriched data back to Excel"""
    print(f"\n💾 Saving results...")

    try:
        df.to_excel(file_path, index=False)
        format_companies_excel(file_path)
        print(f"  ✓ Updated {file_path}")
        return True
    except Exception as e:
        print(f"  ❌ Error saving file: {e}")
        return False


def print_summary(total, successful, errors):
    """Print enrichment summary"""
    print("\n" + "=" * 70)
    print("✅ ENRICHMENT COMPLETE!")
    print("=" * 70)
    print(f"  Companies processed: {total}")
    print(f"  Successfully enriched: {successful}")
    print(f"  Errors: {errors}")
    print("=" * 70 + "\n")


def main():
    """Main execution function"""
    print_header()

    # Validate configuration
    if not Config.validate():
        return 1

    # File paths
    companies_file = os.path.join(os.getcwd(), 'companies.xlsx')

    # Load companies
    df = load_companies(companies_file)
    if df is None:
        return 1

    # Create backup
    create_backup(companies_file)

    # Identify incomplete records
    print("\n🔍 Identifying companies with incomplete data...")
    incomplete_df = identify_incomplete_records(df)

    if len(incomplete_df) == 0:
        print("\n✅ All companies have complete data. Nothing to enrich.")
        return 0

    # Process each incomplete company
    print(f"\n🔬 Enriching {len(incomplete_df)} companies...\n")

    successful = 0
    errors = 0

    for i, (idx, row) in enumerate(incomplete_df.iterrows(), 1):
        company_name = row['Company']
        print(f"[{i}/{len(incomplete_df)}] {company_name}")

        if enrich_company(df, idx, company_name):
            successful += 1
        else:
            errors += 1

        # Small delay to avoid rate limiting
        if i < len(incomplete_df):
            import time
            time.sleep(2)

    # Delete companies with Type = "Unknown"
    df = delete_unknown_companies(df)

    # Save results
    if save_results(df, companies_file):
        print_summary(len(incomplete_df), successful, errors)
        return 0
    else:
        return 1


if __name__ == "__main__":
    sys.exit(main())
