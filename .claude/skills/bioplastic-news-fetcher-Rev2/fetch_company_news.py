#!/usr/bin/env python3
"""
Bioplastic News Fetcher Rev2
Fetches general bioplastic news and matches companies with fuzzy matching
"""

import sys
import os
import json
import logging
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Add parent directory to path to import config
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent))
from config import Config

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bioplastic_news_errors.log'),
        logging.StreamHandler()
    ]
)

def get_iso_week(date_str):
    """Convert date string to ISO week format (e.g., '2025-W43')"""
    try:
        date_obj = pd.to_datetime(date_str)
        iso_year, iso_week, _ = date_obj.isocalendar()
        return f"{iso_year}-W{iso_week:02d}"
    except Exception as e:
        logging.error(f"Error converting date {date_str} to ISO week: {e}")
        return None

def fuzzy_match_company(company_name, known_companies, threshold=85):
    """
    Fuzzy match a company name against a list of known companies
    Returns (matched_name, score) or (None, 0) if no match above threshold
    """
    best_match = None
    best_score = 0

    for known_company in known_companies:
        score = fuzz.ratio(company_name.lower(), known_company.lower())
        if score > best_score:
            best_score = score
            best_match = known_company

    if best_score >= threshold:
        return best_match, best_score
    return None, best_score

def validate_url(url):
    """
    Validate that a URL exists and is accessible
    Returns: (is_valid, status_code)
    Note: Accepts 200 (OK) and 403 (Forbidden - exists but blocks bots)
    """
    try:
        response = requests.head(url, timeout=5, allow_redirects=True)
        # Accept both 200 (accessible) and 403 (exists but blocks automated access)
        is_valid = response.status_code in [200, 403]
        return is_valid, response.status_code
    except:
        try:
            # Try GET if HEAD fails
            response = requests.get(url, timeout=5, allow_redirects=True)
            is_valid = response.status_code in [200, 403]
            return is_valid, response.status_code
        except:
            return False, 0

def search_bioplastic_news(days=7, max_results=20):
    """
    Use Perplexity Research API with search recency to find real bioplastic news
    Returns list of news items extracted from research
    """
    try:
        Config.validate()

        # Calculate date range
        today = datetime.now()
        start_date = today - timedelta(days=days)

        print(f"\n🔍 Searching bioplastic news from {start_date.strftime('%B %d')} to {today.strftime('%B %d, %Y')}...")

        # Construct search query with time constraint
        search_query = f"""
        Find recent news articles about bioplastic and biopolymer companies published between
        {start_date.strftime('%B %d, %Y')} and {today.strftime('%B %d, %Y')}.

        Focus on:
        - Company announcements (plant openings, partnerships, M&A, product launches)
        - Press releases from bioplastic producers and converters
        - Recent industry news with specific company names and dates

        Provide the actual URLs and publication dates for each news item found.
        """

        # Use Perplexity Chat API with research mode (return_citations)
        headers = {
            "Authorization": f"Bearer {Config.PERPLEXITY_API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": "sonar-pro",
            "messages": [
                {
                    "role": "system",
                    "content": "You are a news researcher. Find real, recent news articles with verified URLs and dates."
                },
                {
                    "role": "user",
                    "content": search_query
                }
            ],
            "max_tokens": 3000,
            "temperature": 0.3,
            "return_citations": True,
            "stream": False
        }

        print(f"📡 Calling Perplexity Research API...")
        response = requests.post(
            Config.PERPLEXITY_API_URL,
            headers=headers,
            json=payload,
            timeout=30
        )

        if response.status_code == 200:
            result = response.json()
            content = result['choices'][0]['message']['content']
            citations = result.get('citations', [])

            print(f"✓ Found {len(citations)} cited sources")
            print(f"\n📄 Research findings:\n{content}\n")

            if citations:
                print(f"\n📚 Citations:")
                for idx, url in enumerate(citations[:10], 1):
                    print(f"  [{idx}] {url}")

            # Return both content and citations for extraction
            return {
                'content': content,
                'citations': citations
            }
        else:
            logging.error(f"Research API Error {response.status_code}: {response.text}")
            return None

    except Exception as e:
        logging.error(f"Error searching news: {e}")
        return None

def extract_news_from_results(research_data, days=7):
    """
    Extract structured news from research results
    Returns JSON array with: Company, Publishing Date, Headline, Description, Category, Source URL
    """
    if not research_data or not research_data.get('content'):
        return []

    try:
        Config.validate()

        # Calculate date range
        today = datetime.now()
        start_date = today - timedelta(days=days)

        content = research_data['content']
        citations = research_data.get('citations', [])

        # Build citations list
        citations_text = "\n".join([f"[{idx+1}] {url}" for idx, url in enumerate(citations)])

        # Construct extraction query
        query = f"""
        Based on this research about bioplastic news:

        {content}

        Available citation URLs:
        {citations_text}

        Extract structured news items about bioplastic/biopolymer COMPANIES that match these criteria:
        - From actual companies (producers, converters, compounders, equipment manufacturers)
        - Published between {start_date.strftime('%B %d, %Y')} and {today.strftime('%B %d, %Y')}
        - Categories: Plant Announcement, People Moves, M&A, Product Launch, Partnerships,
          Financial Results, Supply Agreements, Investment & Funding, Certifications

        Return JSON array format:
        [
          {{
            "Company": "company name",
            "PublishingDate": "YYYY-MM-DD",
            "Headline": "headline (max 100 chars)",
            "Description": "50-word summary",
            "Category": "category from list above",
            "SourceURL": "URL from citations list above"
          }}
        ]

        ONLY use URLs from the citations list. Return empty array if no valid news found.
        """

        headers = {
            "Authorization": f"Bearer {Config.PERPLEXITY_API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": Config.DEFAULT_MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": "You extract structured JSON data. Return only valid JSON arrays."
                },
                {
                    "role": "user",
                    "content": query
                }
            ],
            "max_tokens": 2000,
            "temperature": 0.2,
            "stream": False
        }

        print(f"📡 Extracting structured news data...")
        response = requests.post(
            Config.PERPLEXITY_API_URL,
            headers=headers,
            json=payload,
            timeout=30
        )

        if response.status_code == 200:
            result = response.json()
            content = result['choices'][0]['message']['content']

            print(f"✓ Extraction successful")
            print(f"\n📄 Raw response:\n{content}\n")

            # Parse JSON
            content = content.strip()
            if content.startswith('```json'):
                content = content[7:]
            if content.startswith('```'):
                content = content[3:]
            if content.endswith('```'):
                content = content[:-3]
            content = content.strip()

            try:
                news_items = json.loads(content)
                print(f"✓ Successfully parsed {len(news_items)} news items")
                return news_items
            except json.JSONDecodeError as e:
                logging.error(f"Failed to parse JSON: {e}")
                return []
        else:
            logging.error(f"API Error {response.status_code}: {response.text}")
            return []

    except Exception as e:
        logging.error(f"Error extracting news: {e}")
        return []

def fetch_bioplastic_news(days=7, max_items=10, exclude_urls=None):
    """
    Fetch bioplastic news using research-based approach with citations
    Returns JSON array with: Company, Publishing Date, Headline, Description, Category, Source URL
    """
    # Step 1: Research news with citations
    research_data = search_bioplastic_news(days=days)

    if not research_data:
        print("⚠️  No research data found")
        return []

    # Step 2: Extract structured news from research
    news_items = extract_news_from_results(research_data, days=days)

    # Step 3: Filter out excluded URLs
    if exclude_urls and len(exclude_urls) > 0:
        original_count = len(news_items)
        news_items = [item for item in news_items if item.get('SourceURL') not in exclude_urls]
        filtered_count = original_count - len(news_items)
        if filtered_count > 0:
            print(f"  ℹ️  Filtered out {filtered_count} already-seen URLs")

    return news_items[:max_items]  # Limit to max_items

def validate_news_item(item, start_date, end_date):
    """
    Validate a news item for quality and date range
    Returns: (is_valid, reason)
    """
    # List of invalid company patterns (non-companies)
    invalid_patterns = [
        'market', 'industry', 'report', 'insights', 'analysis', 'news',
        'publication', 'association', 'plastics industry', 'biopolymers market',
        'research', 'study', 'survey', 'forecast', 'outlook'
    ]

    company_name = item.get('Company', '').lower()

    # Check if it's a non-company entity
    for pattern in invalid_patterns:
        if pattern in company_name:
            return False, f"Not a company (contains '{pattern}')"

    # Validate date
    try:
        pub_date = pd.to_datetime(item.get('PublishingDate', ''))
        # Normalize all dates to midnight for fair comparison
        pub_date = pub_date.normalize()
        start_date_normalized = pd.Timestamp(start_date).normalize()
        end_date_normalized = pd.Timestamp(end_date).normalize()

        if pub_date < start_date_normalized or pub_date > end_date_normalized:
            return False, f"Date {pub_date.date()} outside range ({start_date_normalized.date()} to {end_date_normalized.date()})"
    except:
        return False, "Invalid date format"

    # Validate required fields
    required_fields = ['Company', 'PublishingDate', 'Headline', 'Description', 'Category', 'SourceURL']
    for field in required_fields:
        if not item.get(field):
            return False, f"Missing required field: {field}"

    # Validate category
    valid_categories = [
        'Plant Announcement', 'People Moves', 'M&A', 'Litigation',
        'Product Launch', 'Partnerships', 'Financial Results',
        'Supply Agreements', 'Investment & Funding', 'Certifications'
    ]
    if item.get('Category') not in valid_categories:
        return False, f"Invalid category: {item.get('Category')}"

    return True, "Valid"

def process_news_items(news_items, companies_df, start_date, end_date):
    """
    Process news items: validate, fuzzy match companies, add new companies if needed
    Returns: processed_items (list of dicts), updated_companies_df
    """
    processed_items = []
    known_companies = companies_df['Company'].tolist()
    new_companies = []

    print(f"\n🔍 Processing {len(news_items)} news items...")

    for item in news_items:
        # Validate news item
        is_valid, reason = validate_news_item(item, start_date, end_date)
        if not is_valid:
            print(f"  ⚠️  Skipping '{item.get('Company', 'Unknown')}': {reason}")
            continue

        # Validate URL exists
        source_url = item.get('SourceURL', '')
        if source_url:
            print(f"  🔗 Validating URL: {source_url[:60]}...")
            url_valid, status_code = validate_url(source_url)
            if not url_valid:
                print(f"  ⚠️  Skipping '{item.get('Company', 'Unknown')}': Invalid URL (status {status_code})")
                continue
            print(f"  ✓ URL valid (status {status_code})")
        company_name = item.get('Company', '')

        # Fuzzy match against known companies
        matched_company, score = fuzzy_match_company(company_name, known_companies)

        if matched_company:
            print(f"  ✓ Matched '{company_name}' to '{matched_company}' (score: {score})")
        else:
            print(f"  🆕 New company found: '{company_name}' (best score: {score})")
            if company_name not in new_companies:
                new_companies.append(company_name)
                # Add to known companies list for subsequent fuzzy matching
                known_companies.append(company_name)

        # Categorize URLs
        source_url = item.get('SourceURL', '')
        company_url = ''
        other_url = ''

        if matched_company:
            # Check if URL matches company webpage
            company_webpage = companies_df[companies_df['Company'] == matched_company]['Webpage'].values
            if len(company_webpage) > 0 and pd.notna(company_webpage[0]) and company_webpage[0]:
                webpage_domain = str(company_webpage[0]).replace('www.', '').lower()
                if webpage_domain in source_url.lower():
                    company_url = source_url
                else:
                    other_url = source_url
            else:
                other_url = source_url
        else:
            other_url = source_url

        # Calculate ISO week
        publishing_date = item.get('PublishingDate', '')
        iso_week = get_iso_week(publishing_date)

        processed_item = {
            'Company': company_name,
            'Company matched': matched_company if matched_company else '',
            'Publishing Date': publishing_date,
            'Headline': item.get('Headline', ''),
            'Description': item.get('Description', ''),
            'Category': item.get('Category', ''),
            'Source URL (company)': company_url,
            'Source URL (other)': other_url,
            'Week': iso_week,
            'Source Skill': 'Perplexity Rev2',
            'Story Generated': 'No'
        }

        processed_items.append(processed_item)

    # Add new companies to companies DataFrame
    if new_companies:
        print(f"\n🆕 Adding {len(new_companies)} new companies to companies.xlsx:")
        for company in new_companies:
            print(f"  - {company}")
            new_row = pd.DataFrame({
                'Company': [company],
                'Type': [''],
                'Webpage': [''],
                'Description': ['']
            })
            companies_df = pd.concat([companies_df, new_row], ignore_index=True)

    return processed_items, companies_df

def format_news_excel(file_path):
    """
    Format Excel file with clickable URLs, proper column widths, and text wrapping
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Define column widths and wrap settings
        column_config = {
            'A': ('ID', 8, False),                    # ID
            'B': ('Company', 20, False),              # Company
            'C': ('Company matched', 20, False),      # Company matched
            'D': ('Publishing Date', 15, False),      # Publishing Date
            'E': ('Headline', 50, True),              # Headline - wrap
            'F': ('Description', 60, True),           # Description - wrap
            'G': ('Category', 20, False),             # Category
            'H': ('Source URL (company)', 40, False), # Source URL (company)
            'I': ('Source URL (other)', 40, False),   # Source URL (other)
            'J': ('Week', 12, False),                 # Week
            'K': ('Source Skill', 20, False),         # Source Skill
            'L': ('Story Generated', 15, False),      # Story Generated
        }

        # Set column widths and wrap text
        for col_letter, (col_name, width, wrap) in column_config.items():
            ws.column_dimensions[col_letter].width = width

            # Apply to all cells in column (skip header)
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if wrap:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Make URLs clickable (columns H and I)
        for row in range(2, ws.max_row + 1):
            # Source URL (company) - column H
            cell_h = ws[f'H{row}']
            if cell_h.value and str(cell_h.value).startswith('http'):
                cell_h.hyperlink = cell_h.value
                cell_h.style = 'Hyperlink'

            # Source URL (other) - column I
            cell_i = ws[f'I{row}']
            if cell_i.value and str(cell_i.value).startswith('http'):
                cell_i.hyperlink = cell_i.value
                cell_i.style = 'Hyperlink'

        wb.save(file_path)
    except Exception as e:
        logging.warning(f"Could not format Excel file: {e}")

def format_companies_excel(file_path):
    """
    Format companies Excel file with clickable URLs and proper widths
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Define column widths and wrap settings
        column_config = {
            'A': ('Company', 25, False),
            'B': ('Type', 20, False),
            'C': ('Country', 15, False),
            'D': ('Webpage', 40, False),
            'E': ('Description', 70, True),
            'F': ('Primary Materials', 50, True),
            'G': ('Market Segments', 50, True),
            'H': ('Status', 12, False),
            'I': ('Publicly Listed', 15, False),
            'J': ('Stock Ticker', 15, False),
            'K': ('Date Added', 15, False),
        }

        # Set column widths and wrap text
        for col_letter, (col_name, width, wrap) in column_config.items():
            ws.column_dimensions[col_letter].width = width

            # Apply to all cells in column (skip header)
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if wrap:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Make Webpage URLs clickable (column D)
        for row in range(2, ws.max_row + 1):
            cell = ws[f'D{row}']
            if cell.value:
                url = str(cell.value).strip()
                # Add https:// if it starts with www. or doesn't have a protocol
                if url and not url.startswith(('http://', 'https://')):
                    if url.startswith('www.'):
                        url = 'https://' + url
                    elif '.' in url:  # Looks like a domain
                        url = 'https://' + url

                if url.startswith('http'):
                    cell.hyperlink = url
                    cell.style = 'Hyperlink'

        wb.save(file_path)
    except Exception as e:
        logging.warning(f"Could not format companies Excel file: {e}")

def manual_review_items(processed_items, auto_approve=False):
    """
    Show each news item to user for manual approval
    Returns: list of approved items
    """
    if not processed_items:
        return []

    # Auto-approve mode (non-interactive)
    if auto_approve or not sys.stdin.isatty():
        print(f"\n" + "=" * 70)
        print(f"📋 AUTO-APPROVE MODE: {len(processed_items)} items")
        print("=" * 70)

        for idx, item in enumerate(processed_items, 1):
            print(f"\n[{idx}/{len(processed_items)}] ✓ AUTO-APPROVED")
            print(f"  Company: {item['Company']}")
            if item['Company matched']:
                print(f"  Matched to: {item['Company matched']}")
            print(f"  Date: {item['Publishing Date']}")
            print(f"  Category: {item['Category']}")
            print(f"  Headline: {item['Headline']}")
            print(f"  Description: {item['Description']}")
            url = item.get('Source URL (company)') or item.get('Source URL (other)', '')
            if url:
                print(f"  URL: {url}")

        print(f"\n" + "=" * 70)
        print(f"✓ All {len(processed_items)} items auto-approved")
        print("=" * 70)
        return processed_items

    # Manual review mode (interactive)
    print(f"\n" + "=" * 70)
    print(f"📋 MANUAL REVIEW: {len(processed_items)} items to review")
    print("=" * 70)

    approved_items = []

    for idx, item in enumerate(processed_items, 1):
        print(f"\n[{idx}/{len(processed_items)}]")
        print(f"  Company: {item['Company']}")
        if item['Company matched']:
            print(f"  Matched to: {item['Company matched']}")
        print(f"  Date: {item['Publishing Date']}")
        print(f"  Category: {item['Category']}")
        print(f"  Headline: {item['Headline']}")
        print(f"  Description: {item['Description']}")

        # Show URL
        url = item.get('Source URL (company)') or item.get('Source URL (other)', '')
        if url:
            print(f"  URL: {url}")

        # Ask for approval
        while True:
            response = input("\n  Approve this item? (y/n/q to quit): ").lower().strip()
            if response == 'y':
                approved_items.append(item)
                print("  ✓ Approved")
                break
            elif response == 'n':
                print("  ✗ Rejected")
                break
            elif response == 'q':
                print("\n⚠️  Review stopped by user")
                print(f"  Approved {len(approved_items)} out of {idx-1} reviewed items")
                return approved_items
            else:
                print("  Please enter 'y' (yes), 'n' (no), or 'q' (quit)")

    print(f"\n" + "=" * 70)
    print(f"✓ Review complete: {len(approved_items)}/{len(processed_items)} items approved")
    print("=" * 70)

    return approved_items

def save_results(processed_items, companies_df, output_file='companies_news.xlsx', companies_file='companies.xlsx'):
    """
    Save processed news items to companies_news.xlsx and updated companies to companies.xlsx
    """
    print(f"\n💾 Saving results...")

    # Save companies (including any new ones)
    companies_df.to_excel(companies_file, index=False)
    format_companies_excel(companies_file)
    print(f"  ✓ Updated {companies_file}")

    # Create new DataFrame with the new structure
    news_df = pd.DataFrame(processed_items)

    # Append to existing companies_news.xlsx if it exists and has data
    if os.path.exists(output_file):
        try:
            existing_df = pd.read_excel(output_file)

            # Add Source Skill column if it doesn't exist
            if 'Source Skill' not in existing_df.columns:
                existing_df['Source Skill'] = 'Unknown'

            # Check if existing file has data and expected columns
            required_columns = ['Source URL (company)', 'Source URL (other)']
            has_valid_data = (len(existing_df) > 0 and
                            all(col in existing_df.columns for col in required_columns))

            if has_valid_data:
                # Deduplicate by URL before appending
                existing_urls = set()
                existing_urls.update(existing_df['Source URL (company)'].dropna().tolist())
                existing_urls.update(existing_df['Source URL (other)'].dropna().tolist())

                # Filter out duplicates from new items
                original_count = len(news_df)
                news_df_filtered = news_df[
                    ~(news_df['Source URL (company)'].isin(existing_urls) |
                      news_df['Source URL (other)'].isin(existing_urls))
                ]
                duplicates_removed = original_count - len(news_df_filtered)

                if duplicates_removed > 0:
                    print(f"  ℹ️  Removed {duplicates_removed} duplicate(s) by URL")

                # Add IDs to new items
                if 'ID' in existing_df.columns:
                    # Get the next ID after the max existing ID
                    max_id = existing_df['ID'].max() if len(existing_df) > 0 else 0
                    news_df_filtered.insert(0, 'ID', range(max_id + 1, max_id + 1 + len(news_df_filtered)))
                else:
                    # No existing IDs, create them for all
                    existing_df.insert(0, 'ID', range(1, len(existing_df) + 1))
                    news_df_filtered.insert(0, 'ID', range(len(existing_df) + 1, len(existing_df) + 1 + len(news_df_filtered)))

                # Combine existing and deduplicated new data
                combined_df = pd.concat([existing_df, news_df_filtered], ignore_index=True)
                combined_df.to_excel(output_file, index=False)
                format_news_excel(output_file)
                print(f"  ✓ Appended {len(news_df_filtered)} news items to {output_file} (total: {len(combined_df)})")
                if len(news_df_filtered) > 0:
                    print(f"  📋 New IDs: #{max_id + 1} to #{max_id + len(news_df_filtered)}")
            else:
                # File exists but is empty or invalid, overwrite it with IDs
                news_df.insert(0, 'ID', range(1, len(news_df) + 1))
                news_df.to_excel(output_file, index=False)
                format_news_excel(output_file)
                print(f"  ✓ Saved {len(processed_items)} news items to {output_file} (overwrote empty file)")
                print(f"  📋 IDs: #1 to #{len(news_df)}")
        except Exception as e:
            logging.error(f"Error appending to existing file: {e}")
            news_df.to_excel(output_file, index=False)
            format_news_excel(output_file)
            print(f"  ✓ Saved {len(processed_items)} news items to {output_file} (new file)")
    else:
        # New file - add IDs starting from 1
        news_df.insert(0, 'ID', range(1, len(news_df) + 1))
        news_df.to_excel(output_file, index=False)
        format_news_excel(output_file)
        print(f"  ✓ Saved {len(processed_items)} news items to {output_file} (new file)")
        print(f"  📋 IDs: #1 to #{len(news_df)}")

    return True

def main():
    """Main execution function"""
    try:
        print("=" * 70)
        print("🌱 BIOPLASTIC NEWS FETCHER REV2")
        print("=" * 70)

        # Determine paths (works from both project root and skill directory)
        if os.path.exists('companies.xlsx'):
            companies_file = 'companies.xlsx'
            news_file = 'companies_news.xlsx'
        else:
            # Running from skill directory, go up to project root
            project_root = Path(__file__).resolve().parent.parent.parent.parent
            companies_file = project_root / 'companies.xlsx'
            news_file = project_root / 'companies_news.xlsx'

        # Load companies
        print("\n📂 Loading companies file...")
        companies_df = pd.read_excel(companies_file)
        print(f"  ✓ Loaded {len(companies_df)} companies")

        # Define date range for validation
        days = 7
        today = datetime.now()
        start_date = today - timedelta(days=days)
        current_iso_week = get_iso_week(today.strftime('%Y-%m-%d'))

        # Load existing news and get URLs from current week
        exclude_urls = []
        if os.path.exists(news_file):
            try:
                existing_news_df = pd.read_excel(news_file)
                # Filter for current week
                current_week_news = existing_news_df[existing_news_df['Week'] == current_iso_week]
                # Collect all URLs from both URL columns
                urls_company = current_week_news['Source URL (company)'].dropna().tolist()
                urls_other = current_week_news['Source URL (other)'].dropna().tolist()
                exclude_urls = list(set(urls_company + urls_other))  # Remove duplicates
                print(f"📋 Found {len(current_week_news)} existing news items for week {current_iso_week}")
            except Exception as e:
                logging.warning(f"Could not load existing news for URL exclusion: {e}")

        # Fetch news
        news_items = fetch_bioplastic_news(days=days, max_items=10, exclude_urls=exclude_urls)

        if not news_items:
            print("\n⚠️  No news items found or API error occurred")
            return

        # Process news items with date validation
        processed_items, updated_companies_df = process_news_items(
            news_items, companies_df, start_date, today
        )

        # Manual review - let user approve each item
        if processed_items:
            approved_items = manual_review_items(processed_items)
        else:
            approved_items = []
            print("\n⚠️  No items to review")

        # Save only approved results
        if approved_items:
            save_results(approved_items, updated_companies_df,
                        output_file=str(news_file), companies_file=str(companies_file))
        else:
            print("\n⚠️  No approved items to save")

        # Summary with category breakdown
        print("\n" + "=" * 70)
        print("✅ REV2 PROCESSING COMPLETE!")
        print("=" * 70)
        print(f"  Approved news items: {len(approved_items)}")
        print(f"  New companies discovered: {len(updated_companies_df) - len(companies_df)}")
        print(f"  Total companies in database: {len(updated_companies_df)}")

        if approved_items:
            print("\n📊 Category breakdown:")
            categories = {}
            for item in approved_items:
                cat = item.get('Category', 'Unknown')
                categories[cat] = categories.get(cat, 0) + 1
            for cat, count in sorted(categories.items()):
                print(f"     {cat}: {count}")

        print("=" * 70)

    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user")
        logging.info("Process interrupted by user")
    except Exception as e:
        logging.error(f"Fatal error: {e}")
        print(f"\n❌ Fatal error: {e}")

if __name__ == "__main__":
    main()
